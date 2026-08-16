"""Route-layer tests for the delivery/dispatch HTTP API.

Why this file exists
--------------------
Before 2026-07-31 the delivery module had 49 tests, all of which imported
service modules directly. **Nothing exercised the route layer**, and that is
exactly where every Critical bug in docs/DELIVERY_AUDIT_2026-07-31.md lived:

  - C-01, the import that made GPS silently return an empty list, was one
    line inside a request handler;
  - C-02/C-03, the normalization and plate-matching failures, were only
    observable in an assembled response;
  - C-04, the total absence of authentication, is a property of routes;
  - C-05's duplicate-vehicle write happened behind an endpoint.

A service-level suite cannot see any of those. Worse, the GPS tests that did
exist asserted the *wrong* input contract, so they passed against a function
that could never work in production — false confidence precisely where the
bugs were (audit T-01, T-02).

These tests drive real HTTP through `app.test_client()` with TTAS mocked, so
they cover the assembled request → service → database → JSON path.
"""
import io
import os
import sqlite3
import sys
import tempfile
import zipfile
from datetime import date, datetime, timedelta
from pathlib import Path
from unittest.mock import patch

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

# app/config.py reads DB_PATH at import time and app/__init__.py runs
# init_db() against it, so both must point somewhere disposable before the
# application package is imported. An absolute path overrides config's
# BASE_DIR join.
_BOOT_FD, _BOOT_DB = tempfile.mkstemp(suffix="-boot.db")
os.close(_BOOT_FD)
os.environ["DB_PATH"] = _BOOT_DB

from app import create_app                                    # noqa: E402
from app.database.migrations import add_vehicle_envelope_columns  # noqa: E402
from services.delivery import plan_service, execution_service  # noqa: E402
from services.delivery.database import init_delivery_tables    # noqa: E402

# A raw TTAS DevList item. Note the plate has no hyphen while the fleet
# stores "50E-18463" — the mismatch that produced audit C-03.
#
# `trktime` is day-first, as TTAS actually writes it. This fixture used ISO
# until 2026-08-01, which is precisely why nothing here noticed that the
# dashboard was reading the date month-first and reporting every vehicle
# ~205 days stale. A fixture in a format production never sends is a test
# asserting a contract that does not exist.
TTAS_PAYLOAD = [{
    "biensoxe": "50E18463",
    "latitude": "10.8500",
    "longitude": "106.6500",
    "speed": "Chạy 42km/h",
    "ad3": "Nổ",
    "trktime": "31/07/2026 09:00:00",
    "driver": "Driver A",
    "devimei": "IMEI-1",
}]


@pytest.fixture(autouse=True)
def isolated_upload_root(tmp_path, monkeypatch):
    """Redirect proof-of-delivery uploads into a per-test temp directory.

    image_service resolves UPLOAD_ROOT from its own file location, so without
    this the suite writes real .jpg files into the repository's DeliveryPlans/
    folder and leaves them there — the existing service tests have been doing
    exactly that, accumulating stray files across runs.
    """
    from services.delivery import image_service

    root = tmp_path / "DeliveryPlans"
    root.mkdir()
    monkeypatch.setattr(image_service, "DATA_ROOT", tmp_path)
    monkeypatch.setattr(image_service, "UPLOAD_ROOT", root)
    yield root


@pytest.fixture(scope="module")
def app():
    application = create_app()
    application.config["TESTING"] = True
    return application


@pytest.fixture
def db(app):
    """A fresh delivery database per test, wired into the app."""
    fd, path = tempfile.mkstemp(suffix=".db")
    os.close(fd)
    init_delivery_tables(path)

    conn = sqlite3.connect(path)
    conn.execute("""
        CREATE TABLE vehicles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            plate_number TEXT NOT NULL UNIQUE,
            vehicle_type TEXT NOT NULL DEFAULT '',
            current_driver TEXT NOT NULL DEFAULT '',
            container_config_id INTEGER DEFAULT NULL
        )
    """)
    # Run the real migration rather than restating its column list here: this
    # fixture hand-writes `vehicles` (the delivery schema doesn't own it), and
    # a duplicated list drifts silently until a query on the new columns blows
    # up in a suite that has nothing to do with them.
    add_vehicle_envelope_columns(conn)
    conn.execute(
        "INSERT INTO vehicles (plate_number, vehicle_type, current_driver) "
        "VALUES ('50E-18463', 'Box Truck', 'Original Driver')"
    )
    conn.commit()
    conn.close()

    app.config["DB_PATH"] = path
    yield path
    os.unlink(path)


@pytest.fixture
def client(app, db):
    """HTTP client. Every endpoint is open — the dispatcher password was
    removed 2026-07-31, so there is no authenticated variant of this."""
    with app.test_client() as c:
        yield c


@pytest.fixture
def plan(db):
    """A confirmed plan with one assignment and three stops, dated *today*.

    The date matters now: correctability is decided per plan-day
    (execution_service.can_revert), so a hard-coded past date would silently
    put every revert test on the refusal path.
    """
    plan_id = plan_service.create_plan(db, "Route Plan", date.today().isoformat())
    plan_service.update_plan(db, plan_id, status="confirmed")
    assignment_id = plan_service.create_assignment(db, plan_id, 1, sequence=1)
    stop_ids = [
        plan_service.create_stop(
            db, assignment_id, i, station_code=f"S{i}", station_name=f"Stop {i}",
            address=f"{i} Main St", lat=10.8 + i / 100, lng=106.6 + i / 100,
            manager_name="Mr T", manager_phone="0900000000",
        )
        for i in (1, 2, 3)
    ]
    return {"plan_id": plan_id, "assignment_id": assignment_id, "stop_ids": stop_ids}


def with_gps(payload=None):
    return patch("services.delivery.routes.fetch_vehicle_data",
                 return_value=(TTAS_PAYLOAD if payload is None else payload, "live", None))


def _ddmm(iso_date):
    """`2026-08-02` → `02_08`, the operator's subfolder date format."""
    y, m, d = str(iso_date)[:10].split("-")
    return f"{d}_{m}"


def _set_driver_override(db_path, assignment_id, name):
    """Name the assignment's driver directly.

    `day_summary` resolves driver_name as COALESCE(override, drivers.name,
    vehicles.current_driver), and the override is the only one of the three
    this suite can set without inventing a driver row.
    """
    conn = sqlite3.connect(db_path)
    try:
        conn.execute(
            "UPDATE vehicle_assignments SET driver_name_override = ? WHERE id = ?",
            (name, assignment_id))
        conn.commit()
    finally:
        conn.close()


def _give_proof(db_path, stop_id):
    """Attach the photos a completion requires, without touching the disk.

    The gate reads delivery_stop_images rather than the filesystem, so rows
    are enough — and a test that had to post real .jpg files just to reach
    'completed' would be testing the upload path all over again.
    """
    conn = sqlite3.connect(db_path)
    for cat in execution_service.PROOF_CATEGORIES:
        conn.execute(
            "INSERT INTO delivery_stop_images (stop_id, category, filename, relative_path) "
            "VALUES (?, ?, ?, ?)",
            (stop_id, cat, f"{cat}.jpg", f"DeliveryPlans/{cat}.jpg"),
        )
    conn.commit()
    conn.close()


def _age_execution(db_path, stop_id, minutes):
    """Push a stop's action timestamps into the past.

    Used to show correctability is *not* governed by elapsed time: the rule
    is the plan's date, so an action hours old on today's plan must still be
    correctable.
    """
    then = (datetime.now() - timedelta(minutes=minutes)).isoformat()
    conn = sqlite3.connect(db_path)
    conn.execute(
        "UPDATE stop_executions SET "
        "  actual_arrival_at = CASE WHEN actual_arrival_at IS NULL THEN NULL ELSE ? END, "
        "  completed_at      = CASE WHEN completed_at      IS NULL THEN NULL ELSE ? END "
        "WHERE stop_id = ?",
        (then, then, stop_id),
    )
    conn.commit()
    conn.close()


# ===========================================================================
# GPS pipeline (audit C-01, C-02, C-03)
# ===========================================================================

class TestAssignmentDriverName:
    """The driver typed in the plan builder has to reach the dispatch page.

    The service suite is blind to this one: the builder POSTs to
    /api/assignments, and the field was being dropped in the route handler,
    not in plan_service. Everything below therefore goes over HTTP.
    """

    def _post_assignment(self, client, plan_id, **extra):
        return client.post("/api/assignments", json={
            "plan_id": plan_id, "vehicle_id": 1, "sequence": 1, **extra,
        })

    def test_posted_name_reaches_the_dispatch_dashboard(self, client, db):
        plan_id = plan_service.create_plan(db, "P", date.today().isoformat())
        plan_service.update_plan(db, plan_id, status="confirmed")
        r = self._post_assignment(client, plan_id, driver_name="Nguyen Van Thay")
        assert r.status_code == 201

        with with_gps():
            body = client.get("/api/execution/dashboard").get_json()
        assert body["assignments"][0]["current_driver"] == "Nguyen Van Thay", \
            "dispatcher sees the vehicle's default instead of who is driving today"

    def test_omitting_the_name_keeps_the_vehicle_default(self, client, db):
        plan_id = plan_service.create_plan(db, "P", date.today().isoformat())
        plan_service.update_plan(db, plan_id, status="confirmed")
        self._post_assignment(client, plan_id)

        with with_gps():
            body = client.get("/api/execution/dashboard").get_json()
        assert body["assignments"][0]["current_driver"] == "Original Driver"

    def test_the_name_comes_back_when_the_plan_is_reopened(self, client, db):
        plan_id = plan_service.create_plan(db, "P", date.today().isoformat())
        self._post_assignment(client, plan_id, driver_name="Nguyen Van Thay")

        body = client.get(f"/api/plans/{plan_id}").get_json()
        assert body["assignments"][0]["driver_name"] == "Nguyen Van Thay"

    def test_the_name_can_be_changed_by_put(self, client, db):
        plan_id = plan_service.create_plan(db, "P", date.today().isoformat())
        aid = self._post_assignment(client, plan_id, driver_name="First").get_json()["id"]

        assert client.put(f"/api/assignments/{aid}",
                          json={"driver_name": "Second"}).status_code == 200
        assert client.get(f"/api/assignments/{aid}").get_json()["driver_name"] == "Second"

    def test_a_typed_name_does_not_become_a_driver_record(self, client, db):
        plan_id = plan_service.create_plan(db, "P", date.today().isoformat())
        self._post_assignment(client, plan_id, driver_name="One Off Guy")

        names = [d["name"] for d in client.get("/api/drivers").get_json()]
        assert "One Off Guy" not in names, "a one-off stand-in must not join the roster"


class TestDashboardGps:
    def test_gps_reaches_the_dashboard(self, client, plan):
        """C-01: `from app import fetch_vehicle_data` raised ImportError on
        every request, was swallowed, and returned an empty list."""
        with with_gps():
            body = client.get("/api/execution/dashboard").get_json()
        assert body["gps_source"] == "live"
        assert body["gps_error"] is None
        assert body["gps_matched"] == 1

    def test_assignment_carries_a_gps_block(self, client, plan):
        with with_gps():
            body = client.get("/api/execution/dashboard").get_json()
        gps = body["assignments"][0]["gps"]
        assert gps is not None, "no GPS attached — the map would render no markers"
        assert gps["device_name"] == "50E18463"
        assert gps["lat"] == pytest.approx(10.85)
        assert gps["lng"] == pytest.approx(106.65)

    def test_telemetry_is_read_from_raw_ttas_keys(self, client, plan):
        """C-02: the normalizer read normalize_vehicle()'s *output* names off
        a *raw* TTAS item, so everything but lat/lng silently defaulted."""
        with with_gps():
            gps = client.get("/api/execution/dashboard").get_json()["assignments"][0]["gps"]
        assert gps["speed_kmh"] == 42.0          # from "speed"
        assert gps["vehicle_status"] == "running"  # derived from the speed phrase
        assert gps["engine_status"] == "Nổ"       # from "ad3"
        assert gps["last_update"] == "31/07/2026 09:00:00"  # from "trktime", raw
        assert gps["driver_name"] == "Driver A"   # from "driver"

    def test_dashboard_carries_a_parsed_timestamp(self, client, plan):
        """The dashboard computes GPS age from this field. Reading the raw
        day-first text in the browser gave 8 January for 1 August."""
        with with_gps():
            gps = client.get("/api/execution/dashboard").get_json()["assignments"][0]["gps"]
        assert gps["last_update_iso"] == "2026-07-31T09:00:00"

    @pytest.mark.parametrize("ttas_plate", [
        "50E-18463", "50E18463", "50E 18463", "50e-18463", "18463",
    ])
    def test_plate_formats_all_match_the_same_vehicle(self, client, plan, ttas_plate):
        """C-03: matching was `.strip().lower()` on both sides."""
        payload = [{**TTAS_PAYLOAD[0], "biensoxe": ttas_plate}]
        with with_gps(payload):
            body = client.get("/api/execution/dashboard").get_json()
        assert body["gps_matched"] == 1, f"{ttas_plate!r} failed to match 50E-18463"

    def test_unknown_plate_does_not_match(self, client, plan):
        payload = [{**TTAS_PAYLOAD[0], "biensoxe": "99Z-00000"}]
        with with_gps(payload):
            body = client.get("/api/execution/dashboard").get_json()
        assert body["gps_matched"] == 0
        assert body["assignments"][0].get("gps") is None

    def test_zero_coordinates_are_reported_as_no_fix(self, client, plan):
        """0,0 is the Gulf of Guinea, not a vehicle position."""
        payload = [{**TTAS_PAYLOAD[0], "latitude": "0", "longitude": "0"}]
        with with_gps(payload):
            gps = client.get("/api/execution/dashboard").get_json()["assignments"][0]["gps"]
        assert gps["lat"] is None and gps["lng"] is None

    def test_malformed_coordinates_do_not_500(self, client, plan):
        payload = [{**TTAS_PAYLOAD[0], "latitude": "", "longitude": "n/a"}]
        with with_gps(payload):
            resp = client.get("/api/execution/dashboard")
        assert resp.status_code == 200

    def test_gps_failure_is_reported_not_hidden(self, client, plan):
        with patch("services.delivery.routes.fetch_vehicle_data",
                   return_value=([], "error", "TTAS unreachable")):
            body = client.get("/api/execution/dashboard").get_json()
        assert body["gps_source"] == "error"
        assert body["gps_error"] == "TTAS unreachable"
        assert body["gps_matched"] == 0


class TestEtaEndpoint:
    def test_eta_returns_legs_for_a_matched_vehicle(self, client, plan):
        with with_gps():
            body = client.get(f"/api/eta?assignment_id={plan['assignment_id']}").get_json()
        assert "error" not in body
        assert len(body["etas"]) == 3
        assert body["gps"]["lat"] == pytest.approx(10.85)

    def test_eta_is_not_double_normalized(self, client, plan):
        """C-02 follow-on: the handler normalized an already-normalized dict,
        whose keys are lat/lng not latitude/longitude, coercing both to 0.0."""
        with with_gps():
            body = client.get(f"/api/eta?assignment_id={plan['assignment_id']}").get_json()
        assert body["gps"]["lat"] != 0.0
        assert body["gps"]["lng"] != 0.0

    def test_eta_reports_missing_gps_cleanly(self, client, plan):
        with patch("services.delivery.routes.fetch_vehicle_data", return_value=([], "error", "x")):
            body = client.get(f"/api/eta?assignment_id={plan['assignment_id']}").get_json()
        assert body["etas"] == []
        assert "not available" in body["error"]

    def test_eta_requires_assignment_id(self, client):
        assert client.get("/api/eta").status_code == 400


# ===========================================================================
# Open access
#
# The dispatcher password (audit C-04) was removed on 2026-07-31 at the
# operator's request: this runs on an internal network and the login step
# was costing dispatchers time on every shift change. What used to be the
# authentication suite is now the inverse regression guard — it fails if a
# gate is ever reintroduced without the frontend being taught about it,
# which is how a dispatcher ends up staring at a silent failed action.
# ===========================================================================

MUTATING_ENDPOINTS = [
    ("post",   "/api/drivers"),
    ("post",   "/api/plans"),
    ("put",    "/api/plans/1"),
    ("delete", "/api/plans/1"),
    ("post",   "/api/plans/batch-delete"),
    ("post",   "/api/plans/clear"),
    ("post",   "/api/plans/1/confirm"),
    ("post",   "/api/plans/import/parse"),
    ("post",   "/api/plans/import/save"),
    ("post",   "/api/assignments"),
    ("put",    "/api/assignments/1"),
    ("delete", "/api/assignments/1"),
    ("post",   "/api/stops"),
    ("put",    "/api/stops/1"),
    ("delete", "/api/stops/1"),
    ("post",   "/api/stops/1/skip"),
    ("post",   "/api/stops/1/cancel"),
    ("post",   "/api/stops/reorder"),
    ("post",   "/api/stops/insert"),
    ("post",   "/api/execution/advance"),
    ("post",   "/api/execution/revert"),
    ("post",   "/api/stops/1/images"),
    ("delete", "/api/images/1"),
]

READ_ENDPOINTS = [
    "/api/drivers",
    "/api/plans",
    "/api/assignments",
    "/api/execution/dashboard",
]


class TestOpenAccess:
    @pytest.mark.parametrize("method,path", MUTATING_ENDPOINTS)
    def test_mutating_endpoint_needs_no_session(self, client, method, path):
        """A 401 or 503 here means a gate came back. Anything else — 200, 400
        for a bad body, 404 for the id that does not exist — is fine; this
        test is about reachability, not about each endpoint's contract."""
        resp = getattr(client, method)(path, json={})
        assert resp.status_code not in (401, 403, 503), (
            f"{method.upper()} {path} returned {resp.status_code} — it is gated again"
        )

    @pytest.mark.parametrize("path", READ_ENDPOINTS)
    def test_read_endpoint_stays_open(self, client, path):
        with with_gps():
            assert getattr(client, "get")(path).status_code == 200

    def test_clear_plans_needs_no_session(self, client, db, plan):
        """The most destructive endpoint: cascade-deletes everything. It is
        deliberately reachable — the confirm dialog in the UI is the only
        thing standing in front of it."""
        assert client.post("/api/plans/clear").status_code == 200
        assert plan_service.list_plans(db) == []

    def test_login_route_is_gone(self, client):
        assert client.get("/login").status_code == 404


# ===========================================================================
# Stop execution lifecycle (audit C-07, C-09, C-06b)
# ===========================================================================

class TestExecutionLifecycle:
    def test_full_advance_progression(self, client, db, plan):
        stop_id = plan["stop_ids"][0]

        r = client.post("/api/execution/advance",
                             json={"stop_id": stop_id, "expected_status": "planned"})
        assert r.status_code == 200 and r.get_json()["status"] == "advanced"

        _give_proof(db, stop_id)
        r = client.post("/api/execution/advance",
                             json={"stop_id": stop_id, "expected_status": "arrived"})
        assert r.status_code == 200 and r.get_json()["status"] == "completed"

        e = execution_service.get_stop_execution(db, stop_id)
        assert e["actual_arrival_at"] and e["actual_departure_at"]

    def test_double_tap_cannot_skip_arrived(self, client, db, plan):
        """C-07: two taps took a stop planned -> arrived -> completed, marking
        it delivered with arrival and departure in the same second."""
        stop_id = plan["stop_ids"][0]
        body = {"stop_id": stop_id, "expected_status": "planned"}

        assert client.post("/api/execution/advance", json=body).status_code == 200
        second = client.post("/api/execution/advance", json=body)

        assert second.status_code == 409
        assert second.get_json()["conflict"] is True
        e = execution_service.get_stop_execution(db, stop_id)
        assert e["status"] == "arrived"
        assert e["actual_departure_at"] is None

    def test_advance_without_token_still_works(self, client, db, plan):
        r = client.post("/api/execution/advance", json={"stop_id": plan["stop_ids"][0]})
        assert r.status_code == 200

    def test_advance_requires_stop_id(self, client):
        assert client.post("/api/execution/advance", json={}).status_code == 400

    def test_skip_and_cancel(self, client, db, plan):
        skip_id, cancel_id = plan["stop_ids"][0], plan["stop_ids"][1]

        assert client.post(f"/api/stops/{skip_id}/skip",
                                json={"reason": "gate locked"}).status_code == 200
        assert client.post(f"/api/stops/{cancel_id}/cancel",
                                json={"reason": "customer closed"}).status_code == 200

        assert execution_service.get_stop_execution(db, skip_id)["skip_reason"] == "gate locked"
        assert execution_service.get_stop_execution(db, cancel_id)["cancel_reason"] == "customer closed"

    def test_current_stop_advances_to_the_next(self, client, db, plan):
        first, second = plan["stop_ids"][0], plan["stop_ids"][1]
        client.post(f"/api/stops/{first}/skip", json={"reason": "x"})

        current = client.get(
            f"/api/execution/current?assignment_id={plan['assignment_id']}"
        ).get_json()
        assert current["id"] == second

    def test_plan_auto_completes_when_every_stop_is_terminal(self, client, db, plan):
        for stop_id in plan["stop_ids"]:
            client.post(f"/api/stops/{stop_id}/skip", json={"reason": "x"})
        assert plan_service.get_plan(db, plan["plan_id"])["status"] == "completed"

    def test_progress_endpoint(self, client, db, plan):
        client.post(f"/api/stops/{plan['stop_ids'][0]}/skip", json={"reason": "x"})
        prog = client.get(
            f"/api/execution/progress?assignment_id={plan['assignment_id']}"
        ).get_json()
        assert (prog["total"], prog["completed"], prog["remaining"]) == (3, 1, 2)

    def test_empty_assignment_reports_zero_not_one(self, client, db, plan):
        """C-09: `total = sum(...) or 1` made an empty assignment claim it had
        one remaining stop, sending a dispatcher after nothing."""
        empty_id = plan_service.create_assignment(db, plan["plan_id"], 1, sequence=2)
        with with_gps():
            body = client.get("/api/execution/dashboard").get_json()
        entry = next(a for a in body["assignments"] if a["assignment_id"] == empty_id)
        assert entry["progress"]["total"] == 0
        assert entry["progress"]["remaining"] == 0


class TestRevertEndpoint:
    """Undo for a mis-tapped Advance/Skip/Cancel.

    Route-layer coverage matters here for the same reason it did for advance:
    the guard that makes revert safe is the `expected_status` token, which
    only exists in the request body, and the `can_revert` flag the button is
    drawn from is assembled in a response — neither is visible from the
    service suite.
    """

    def _advance(self, client, stop_id, expected):
        return client.post("/api/execution/advance",
                           json={"stop_id": stop_id, "expected_status": expected})

    def test_revert_undoes_an_accidental_advance(self, client, db, plan):
        stop_id = plan["stop_ids"][0]
        self._advance(client, stop_id, "planned")

        r = client.post("/api/execution/revert",
                        json={"stop_id": stop_id, "expected_status": "arrived"})

        assert r.status_code == 200 and r.get_json()["status"] == "planned"
        e = execution_service.get_stop_execution(db, stop_id)
        assert e["status"] == "planned"
        assert e["actual_arrival_at"] is None, "an un-arrived stop kept its arrival time"

    def test_revert_from_completed_restores_arrived_not_planned(self, client, db, plan):
        """One step back, not all the way. The driver really did arrive; only
        the second tap was the mistake."""
        stop_id = plan["stop_ids"][0]
        self._advance(client, stop_id, "planned")
        _give_proof(db, stop_id)
        self._advance(client, stop_id, "arrived")

        r = client.post("/api/execution/revert",
                        json={"stop_id": stop_id, "expected_status": "completed"})

        assert r.status_code == 200 and r.get_json()["status"] == "arrived"
        e = execution_service.get_stop_execution(db, stop_id)
        assert e["status"] == "arrived"
        assert e["actual_arrival_at"] is not None
        assert e["actual_departure_at"] is None and e["completed_at"] is None

    def test_reverted_stop_becomes_current_again(self, client, db, plan):
        """The point of the feature: a mis-advanced stop moves the dashboard
        on to the next one, and reverting has to move it back."""
        first, second = plan["stop_ids"][0], plan["stop_ids"][1]
        client.post(f"/api/stops/{first}/skip", json={"reason": "mis-tap"})
        url = f"/api/execution/current?assignment_id={plan['assignment_id']}"
        assert client.get(url).get_json()["id"] == second

        client.post("/api/execution/revert",
                    json={"stop_id": first, "expected_status": "skipped"})

        assert client.get(url).get_json()["id"] == first

    def test_revert_of_skip_clears_the_reason(self, client, db, plan):
        stop_id = plan["stop_ids"][0]
        client.post(f"/api/stops/{stop_id}/skip", json={"reason": "gate locked"})

        client.post("/api/execution/revert", json={"stop_id": stop_id})

        e = execution_service.get_stop_execution(db, stop_id)
        assert e["status"] == "planned"
        assert e["skip_reason"] == ""
        assert e["completed_at"] is None

    def test_revert_of_a_skip_after_arrival_returns_to_arrived(self, client, db, plan):
        """A stop skipped once the driver was already there has a real arrival
        time. Sending it back to 'planned' would either strand that timestamp
        or destroy it."""
        stop_id = plan["stop_ids"][0]
        self._advance(client, stop_id, "planned")
        client.post(f"/api/stops/{stop_id}/skip", json={"reason": "nobody home"})

        r = client.post("/api/execution/revert",
                        json={"stop_id": stop_id, "expected_status": "skipped"})

        assert r.get_json()["status"] == "arrived"
        assert execution_service.get_stop_execution(db, stop_id)["actual_arrival_at"] is not None

    def test_revert_reopens_an_auto_completed_plan(self, client, db, plan):
        """_maybe_complete_plan closed the plan on the last stop; undoing that
        stop has to bring it back into the dashboard's active view, or the
        dispatcher can no longer see the vehicle they just corrected."""
        for stop_id in plan["stop_ids"]:
            client.post(f"/api/stops/{stop_id}/skip", json={"reason": "x"})
        assert plan_service.get_plan(db, plan["plan_id"])["status"] == "completed"

        client.post("/api/execution/revert", json={"stop_id": plan["stop_ids"][2]})

        assert plan_service.get_plan(db, plan["plan_id"])["status"] == "executing"

    def test_stale_token_is_refused_as_a_conflict(self, client, db, plan):
        """Same guard as advance: a Revert button rendered before someone else
        moved the stop must not act on the status it can no longer see."""
        stop_id = plan["stop_ids"][0]
        self._advance(client, stop_id, "planned")
        _give_proof(db, stop_id)
        self._advance(client, stop_id, "arrived")

        r = client.post("/api/execution/revert",
                        json={"stop_id": stop_id, "expected_status": "arrived"})

        assert r.status_code == 409 and r.get_json()["conflict"] is True
        assert execution_service.get_stop_execution(db, stop_id)["status"] == "completed"

    def test_double_tapped_undo_does_not_step_back_twice(self, client, db, plan):
        stop_id = plan["stop_ids"][0]
        self._advance(client, stop_id, "planned")
        _give_proof(db, stop_id)
        self._advance(client, stop_id, "arrived")
        body = {"stop_id": stop_id, "expected_status": "completed"}

        assert client.post("/api/execution/revert", json=body).status_code == 200
        assert client.post("/api/execution/revert", json=body).status_code == 409
        assert execution_service.get_stop_execution(db, stop_id)["status"] == "arrived"

    def test_a_planned_stop_has_nothing_to_revert(self, client, plan):
        r = client.post("/api/execution/revert", json={"stop_id": plan["stop_ids"][0]})
        assert r.status_code == 400
        assert "Cannot revert" in r.get_json()["error"]

    def test_revert_requires_stop_id(self, client):
        assert client.post("/api/execution/revert", json={}).status_code == 400

    def test_a_closed_days_plan_is_refused(self, client, db, plan):
        """The button is gone by then, but the endpoint is open and a page
        left up overnight could still post — so the day rule is enforced
        here, not only in the markup."""
        stop_id = plan["stop_ids"][0]
        self._advance(client, stop_id, "planned")
        plan_service.update_plan(db, plan["plan_id"],
                                 plan_date=(date.today() - timedelta(days=1)).isoformat())

        r = client.post("/api/execution/revert", json={"stop_id": stop_id})

        assert r.status_code == 400
        assert "date has passed" in r.get_json()["error"]
        assert execution_service.get_stop_execution(db, stop_id)["status"] == "arrived"

    def test_an_hours_old_action_on_todays_plan_is_still_correctable(self, client, db, plan):
        """The rule that replaced the 15-minute window."""
        stop_id = plan["stop_ids"][0]
        self._advance(client, stop_id, "planned")
        _age_execution(db, stop_id, minutes=8 * 60)

        assert client.post("/api/execution/revert", json={"stop_id": stop_id}).status_code == 200

    def test_stops_response_carries_can_revert(self, client, db, plan):
        """The dashboard draws its Revert button from this flag alone."""
        stop_id = plan["stop_ids"][0]
        self._advance(client, stop_id, "planned")

        stops = client.get(f"/api/stops?assignment_id={plan['assignment_id']}").get_json()
        by_id = {s["id"]: s for s in stops}

        assert by_id[stop_id]["can_revert"] is True
        assert by_id[plan["stop_ids"][1]]["can_revert"] is False, \
            "an untouched stop offered an undo for something that never happened"

    def test_can_revert_closes_with_the_plan_day(self, client, db, plan):
        stop_id = plan["stop_ids"][0]
        self._advance(client, stop_id, "planned")
        plan_service.update_plan(db, plan["plan_id"],
                                 plan_date=(date.today() - timedelta(days=1)).isoformat())

        stops = client.get(f"/api/stops?assignment_id={plan['assignment_id']}").get_json()

        assert next(s for s in stops if s["id"] == stop_id)["can_revert"] is False


class TestProofGateEndpoint:
    """The completion gate as the dashboard meets it. The distinct 422 and
    `proof_required` flag are what let the UI offer an override instead of
    pattern-matching the message text."""

    def _arrived(self, client, plan):
        stop_id = plan["stop_ids"][0]
        client.post("/api/execution/advance", json={"stop_id": stop_id})
        return stop_id

    def test_completion_without_proof_is_422_not_400(self, client, db, plan):
        stop_id = self._arrived(client, plan)

        r = client.post("/api/execution/advance",
                        json={"stop_id": stop_id, "expected_status": "arrived"})
        body = r.get_json()

        assert r.status_code == 422, "400 would read as a malformed request"
        assert body["proof_required"] is True
        assert body["missing"] == ["unload", "door"]

    def test_the_message_names_the_missing_photo(self, client, db, plan):
        stop_id = self._arrived(client, plan)
        _give_proof(db, stop_id)
        conn = sqlite3.connect(db)
        conn.execute("DELETE FROM delivery_stop_images WHERE stop_id = ? AND category = 'door'",
                     (stop_id,))
        conn.commit()
        conn.close()

        body = client.post("/api/execution/advance", json={"stop_id": stop_id}).get_json()

        assert body["missing"] == ["door"]
        assert "locked door" in body["error"]

    def test_the_stop_does_not_move(self, client, db, plan):
        stop_id = self._arrived(client, plan)
        client.post("/api/execution/advance", json={"stop_id": stop_id})

        assert execution_service.get_stop_execution(db, stop_id)["status"] == "arrived"
        assert execution_service.get_stop_execution(db, stop_id)["completed_at"] is None

    def test_an_override_in_the_body_completes_it(self, client, db, plan):
        stop_id = self._arrived(client, plan)

        r = client.post("/api/execution/advance", json={
            "stop_id": stop_id,
            "expected_status": "arrived",
            "override_reason": "phone battery died",
        })

        assert r.status_code == 200 and r.get_json()["status"] == "completed"

    def test_the_override_reason_reaches_the_history_endpoint(self, client, db, plan):
        stop_id = self._arrived(client, plan)
        client.post("/api/execution/advance",
                    json={"stop_id": stop_id, "override_reason": "phone battery died"})

        events = client.get(f"/api/stops/{stop_id}/history").get_json()

        assert events[-1]["reason"] == "phone battery died"

    def test_uploading_both_photos_unblocks_the_normal_path(self, client, db, plan):
        """End to end through the real upload endpoint, not injected rows."""
        stop_id = self._arrived(client, plan)
        assert _upload(client, stop_id, "goods.jpg", category="unload").status_code == 201
        assert _upload(client, stop_id, "door.jpg", category="door").status_code == 201

        r = client.post("/api/execution/advance", json={"stop_id": stop_id})

        assert r.status_code == 200 and r.get_json()["status"] == "completed"


class TestDayExport:
    """The end-of-day handover.

    The photos are already on disk, organised the way they were *written*
    (year/month/day/plate/station/category). The operator hands over a
    different shape entirely, so these assert the ZIP's structure rather
    than merely that a ZIP came back.
    """

    def _zip(self, client, date_str, name="2_8_BacLieuGiaRai", loading_date=None):
        params = f"date={date_str}&name={name}"
        if loading_date:
            params += f"&loading_date={loading_date}"
        resp = client.get(f"/api/export/day.zip?{params}")
        assert resp.status_code == 200
        assert resp.mimetype == "application/zip"
        return zipfile.ZipFile(io.BytesIO(resp.data))

    def _files(self, zf):
        """Just the photos. Since 2026-08-10 the ZIP also carries empty folder
        entries, and a count of `namelist()` silently includes them."""
        return [n for n in zf.namelist() if not n.endswith("/")]

    def _day_upload(self, client, date_str, category, filename="x.jpg", label=""):
        return client.post("/api/export/day-images", data={
            "file": (io.BytesIO(b"bytes"), filename),
            "date": date_str,
            "category": category,
            "label": label,
        }, content_type="multipart/form-data")

    def test_summary_lists_drivers_and_missing_proof(self, client, db, plan):
        today = date.today().isoformat()
        _upload(client, plan["stop_ids"][0], "a.jpg", category="unload")

        body = client.get(f"/api/export/summary?date={today}").get_json()

        assert body["stop_count"] == 3
        assert body["incomplete_count"] == 3, "all three still lack at least one photo"
        stop = body["drivers"][0]["stops"][0]
        assert stop["missing"] == ["door"]

    def test_summary_reports_a_waived_completion(self, client, db, plan):
        stop_id = plan["stop_ids"][0]
        client.post("/api/execution/advance", json={"stop_id": stop_id})
        client.post("/api/execution/advance",
                    json={"stop_id": stop_id, "override_reason": "phone battery died"})

        body = client.get(f"/api/export/summary?date={date.today().isoformat()}").get_json()

        stop = next(s for d in body["drivers"] for s in d["stops"] if s["stop_id"] == stop_id)
        assert stop["override_reason"] == "phone battery died"

    def test_summary_requires_a_date(self, client):
        assert client.get("/api/export/summary").status_code == 400

    def test_stop_photos_are_filed_by_driver_then_station(self, client, db, plan):
        today = date.today().isoformat()
        _upload(client, plan["stop_ids"][0], "goods.jpg", category="unload")
        _upload(client, plan["stop_ids"][0], "door.jpg", category="door")

        files = self._files(self._zip(client, today))

        # Driver folder is "<name>_<5-digit plate serial>"; the fixture
        # vehicle is 50E-18463 with driver "Original Driver".
        expected_dir = f"2_8_BacLieuGiaRai/OriginalDriver_18463/HinhGiaoHang_{_ddmm(today)}/S1/"
        assert sum(1 for n in files if n.startswith(expected_dir)) == 2, files

    def test_a_photo_that_is_not_proof_is_left_out(self, client, db, plan):
        """An 'extra' shot is not evidence of anything and must not be filed
        alongside the two categories that are."""
        today = date.today().isoformat()
        _upload(client, plan["stop_ids"][0], "random.jpg", category="extra")

        files = self._files(self._zip(client, today))

        # The stop's folder is still created — it is a checklist — but nothing
        # may be filed inside it.
        assert not any("HinhGiaoHang" in n for n in files), files

    def test_loading_photos_are_filed_under_their_driver(self, client, db, plan):
        """One flat folder until 2026-08-10, briefly its own top-level folder
        with driver subfolders, and now inside the driver — the shape of the
        tree the operator handed over. Not split by stop: loading happens in
        one pass as the truck is filled, and a station per shot would be a tap
        for a distinction nobody reads afterwards."""
        today = date.today().isoformat()
        yesterday = (date.today() - timedelta(days=1)).isoformat()
        assert self._day_upload(client, today, "loading", "load1.jpg",
                                label="OriginalDriver_18463").status_code == 201
        assert self._day_upload(client, today, "loading", "load2.jpg",
                                label="OriginalDriver_18463").status_code == 201

        names = self._zip(client, today).namelist()

        folder = (f"2_8_BacLieuGiaRai/OriginalDriver_18463/"
                  f"HinhNhanHang_{_ddmm(yesterday)}/")
        assert f"{folder}load1.jpg" in names, names
        assert f"{folder}load2.jpg" in names, names
        # Loose in the folder — no station level beneath it.
        photos = [n for n in names if n.startswith(folder) and not n.endswith("/")]
        assert all("/" not in n[len(folder):] for n in photos), photos

    def test_loading_photos_of_different_drivers_do_not_share_a_folder(self, client, db, plan):
        today = date.today().isoformat()
        yesterday = (date.today() - timedelta(days=1)).isoformat()
        self._day_upload(client, today, "loading", "mine.jpg", label="OriginalDriver_18463")
        self._day_upload(client, today, "loading", "theirs.jpg", label="NguyenVanA_51234")

        names = self._zip(client, today).namelist()

        ddmm = _ddmm(yesterday)
        assert (f"2_8_BacLieuGiaRai/OriginalDriver_18463/"
                f"HinhNhanHang_{ddmm}/mine.jpg") in names, names
        assert (f"2_8_BacLieuGiaRai/NguyenVanA_51234/"
                f"HinhNhanHang_{ddmm}/theirs.jpg") in names, names

    def test_an_unlabelled_loading_photo_gets_the_fallback_folder(self, client, db, plan):
        """Every `loading` row written before the driver picker existed has a
        blank label, so this is the normal case for historical dates."""
        today = date.today().isoformat()
        yesterday = (date.today() - timedelta(days=1)).isoformat()
        self._day_upload(client, today, "loading", "old.jpg")

        names = self._zip(client, today).namelist()

        assert (f"2_8_BacLieuGiaRai/KhongRoTaiXe/"
                f"HinhNhanHang_{_ddmm(yesterday)}/old.jpg") in names, names

    def test_a_drivers_two_photo_folders_sit_side_by_side(self, client, db, plan):
        """The picker stores the folder name rather than the driver's name so
        these agree without a lookup. If they ever diverge, the operator opens
        two differently-named folders for one driver."""
        today = date.today().isoformat()
        yesterday = (date.today() - timedelta(days=1)).isoformat()
        _upload(client, plan["stop_ids"][0], "goods.jpg", category="unload")
        self._day_upload(client, today, "loading", "load.jpg",
                         label="OriginalDriver_18463")

        names = self._zip(client, today).namelist()

        driver = "2_8_BacLieuGiaRai/OriginalDriver_18463"
        assert f"{driver}/HinhNhanHang_{_ddmm(yesterday)}/load.jpg" in names, names
        assert any(n.startswith(f"{driver}/HinhGiaoHang_{_ddmm(today)}/S1/")
                   and n.endswith(".jpg") for n in names), names

    def test_a_stop_with_no_photos_still_gets_its_folder(self, client, db, plan):
        """The stop folders are a checklist. Before 2026-08-10 an unphotographed
        stop simply had no folder, so a missed stop and a stop that does not
        exist looked identical in the handover."""
        today = date.today().isoformat()
        _upload(client, plan["stop_ids"][0], "goods.jpg", category="unload")

        names = self._zip(client, today).namelist()

        delivery = f"2_8_BacLieuGiaRai/OriginalDriver_18463/HinhGiaoHang_{_ddmm(today)}"
        # The fixture plan has three stops; only S1 was photographed.
        for station in ("S1", "S2", "S3"):
            assert f"{delivery}/{station}/" in names, names

    def test_empty_folders_are_marked_as_directories(self, client, db, plan):
        """Windows Explorer reads the MS-DOS directory flag, not the Unix mode.
        Without it these unpack as zero-byte *files* named after the stop."""
        today = date.today().isoformat()

        zf = self._zip(client, today)
        info = zf.getinfo(
            f"2_8_BacLieuGiaRai/OriginalDriver_18463/HinhGiaoHang_{_ddmm(today)}/S1/")

        assert info.is_dir()
        assert info.external_attr & 0x10, "MS-DOS directory flag not set"

    def test_a_driver_with_nothing_shot_still_gets_both_folders(self, client, db, plan):
        today = date.today().isoformat()
        yesterday = (date.today() - timedelta(days=1)).isoformat()

        names = self._zip(client, today).namelist()

        driver = "2_8_BacLieuGiaRai/OriginalDriver_18463"
        assert f"{driver}/HinhNhanHang_{_ddmm(yesterday)}/" in names, names
        assert f"{driver}/HinhGiaoHang_{_ddmm(today)}/" in names, names

    def test_the_loading_folder_defaults_to_the_day_before(self, client, db, plan):
        today = date.today().isoformat()
        self._day_upload(client, today, "loading", "load.jpg")

        names = self._zip(client, today).namelist()

        yesterday = (date.today() - timedelta(days=1)).isoformat()
        assert any(f"HinhNhanHang_{_ddmm(yesterday)}/" in n for n in names), names

    def test_the_loading_date_can_be_overridden(self, client, db, plan):
        today = date.today().isoformat()
        self._day_upload(client, today, "loading", "load.jpg")

        names = self._zip(client, today, loading_date="2026-07-20").namelist()

        assert any("HinhNhanHang_20_07/" in n for n in names), names

    def test_empty_container_photos_carry_the_driver_name(self, client, db, plan):
        today = date.today().isoformat()
        self._day_upload(client, today, "empty_container", "truck.jpg",
                         label="Huỳnh Quốc Trọng")

        files = self._files(self._zip(client, today))

        container = [n for n in files if "HinhThungTrong/" in n]
        assert len(container) == 1
        # Accents stripped and words run together, matching the operator's
        # existing folders — and đ/Đ handled, which NFD alone does not.
        assert "HuynhQuocTrong_" in container[0], container[0]

    def _manifest_rows(self, zf, root="2_8_BacLieuGiaRai"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(zf.read(f"{root}/manifest.xlsx")))
        return [[c.value for c in row] for row in wb.active.iter_rows()]

    def test_every_zip_carries_a_manifest(self, client, db, plan):
        today = date.today().isoformat()
        _upload(client, plan["stop_ids"][0], "goods.jpg", category="unload")

        rows = self._manifest_rows(self._zip(client, today))

        assert "station_code" in rows[0]
        flat = [str(v) for r in rows[1:] for v in r]
        assert "S1" in flat
        assert "door" in flat, "the missing photo must be recorded, not just absent"

    def test_the_manifest_keeps_vietnamese_intact(self, client, db, plan):
        """A CSV until 2026-08-10, and Excel read it in the machine's ANSI
        codepage — `Huỳnh Quốc Trọng` arrived as `Huá»³nh Quá»‘c Trá»ng`. The
        bytes were always fine; .xlsx declares its encoding instead of leaving
        the reader to guess."""
        today = date.today().isoformat()
        name = "Huỳnh Quốc Trọng"
        _set_driver_override(db, plan["assignment_id"], name)

        rows = self._manifest_rows(self._zip(client, today))

        assert any(name in [str(v) for v in r] for r in rows[1:]), rows

    def test_the_manifest_is_a_real_workbook(self, client, db, plan):
        """Guards against it quietly reverting to CSV bytes under an .xlsx
        name, which Excel reports as a corrupt file rather than mojibake."""
        zf = self._zip(client, date.today().isoformat())
        blob = zf.read("2_8_BacLieuGiaRai/manifest.xlsx")

        assert blob[:2] == b"PK", "not a zip container, so not an xlsx"

    def test_the_typed_folder_name_cannot_escape_the_zip(self, client, db, plan):
        """It is free text from a form and becomes a path. S-04 all over
        again if it were trusted."""
        today = date.today().isoformat()
        names = self._zip(client, today, name="../../etc").namelist()

        assert not any(n.startswith("..") or n.startswith("/") for n in names), names

    def test_day_image_upload_rejects_an_unknown_category(self, client, plan):
        r = self._day_upload(client, date.today().isoformat(), "not_a_category")
        assert r.status_code == 400
        assert "Unknown category" in r.get_json()["error"]

    def test_day_image_upload_rejects_a_bad_date(self, client, plan):
        assert self._day_upload(client, "not-a-date", "loading").status_code == 400

    def test_day_images_can_be_listed_and_removed(self, client, db, plan):
        today = date.today().isoformat()
        image_id = self._day_upload(client, today, "loading", "l.jpg").get_json()["id"]

        listed = client.get(f"/api/export/day-images?date={today}&category=loading").get_json()
        assert [i["id"] for i in listed] == [image_id]

        assert client.delete(f"/api/export/day-images/{image_id}").status_code == 200
        assert client.get(f"/api/export/day-images?date={today}").get_json() == []

    def test_removing_a_missing_day_image_404s(self, client, plan):
        assert client.delete("/api/export/day-images/99999").status_code == 404

    def test_a_day_with_nothing_planned_still_exports(self, client, db, plan):
        """An empty ZIP with a manifest beats an error at 6pm. No drivers means
        no driver folders — only the container folder, which is not per-driver."""
        names = self._zip(client, "2026-01-01").namelist()
        assert sorted(names) == [
            "2_8_BacLieuGiaRai/HinhThungTrong/",
            "2_8_BacLieuGiaRai/manifest.xlsx",
        ]


class TestStopHistoryEndpoint:
    """The stored phase log, as the dashboard panel reads it."""

    def test_history_reads_oldest_first(self, client, db, plan):
        stop_id = plan["stop_ids"][0]
        client.post("/api/execution/advance", json={"stop_id": stop_id})
        _give_proof(db, stop_id)
        client.post("/api/execution/advance", json={"stop_id": stop_id})
        client.post("/api/execution/revert", json={"stop_id": stop_id})

        events = client.get(f"/api/stops/{stop_id}/history").get_json()

        assert [(e["from_status"], e["to_status"], e["action"]) for e in events] == [
            ("planned", "arrived", "advance"),
            ("arrived", "completed", "advance"),
            ("completed", "arrived", "revert"),
        ]

    def test_a_cancel_reason_survives_into_the_log(self, client, db, plan):
        """The reason is cleared off the execution row by a revert, so the
        log is the only place it continues to exist."""
        stop_id = plan["stop_ids"][0]
        client.post(f"/api/stops/{stop_id}/cancel", json={"reason": "customer closed"})
        client.post("/api/execution/revert", json={"stop_id": stop_id})

        events = client.get(f"/api/stops/{stop_id}/history").get_json()

        assert execution_service.get_stop_execution(db, stop_id)["cancel_reason"] == ""
        assert events[0]["reason"] == "customer closed"

    def test_an_untouched_stop_has_an_empty_log(self, client, plan):
        assert client.get(f"/api/stops/{plan['stop_ids'][0]}/history").get_json() == []

    def test_a_missing_stop_returns_an_empty_log_not_a_500(self, client):
        r = client.get("/api/stops/99999/history")
        assert r.status_code == 200 and r.get_json() == []


class TestReorderValidation:
    """C-06b: any list was accepted and applied stop-by-stop."""

    def test_full_reorder_succeeds(self, client, db, plan):
        a, b, c = plan["stop_ids"]
        r = client.post("/api/stops/reorder",
                             json={"assignment_id": plan["assignment_id"], "stop_ids": [c, a, b]})
        assert r.status_code == 200
        assert [s["id"] for s in plan_service.list_stops(db, plan["assignment_id"])] == [c, a, b]

    def test_partial_list_is_rejected(self, client, db, plan):
        a, b, _ = plan["stop_ids"]
        r = client.post("/api/stops/reorder",
                             json={"assignment_id": plan["assignment_id"], "stop_ids": [b, a]})
        assert r.status_code == 400
        assert "missing" in r.get_json()["error"]

        seqs = [s["execution_sequence"] for s in plan_service.list_stops(db, plan["assignment_id"])]
        assert len(seqs) == len(set(seqs)), f"duplicate execution_sequence: {seqs}"

    def test_foreign_stop_ids_are_rejected(self, client, db, plan):
        other = plan_service.create_assignment(db, plan["plan_id"], 1, sequence=9)
        foreign = plan_service.create_stop(db, other, 1, station_code="X", lat=10.8, lng=106.6)
        r = client.post("/api/stops/reorder",
                             json={"assignment_id": plan["assignment_id"], "stop_ids": [foreign]})
        assert r.status_code == 400
        assert "not in this assignment" in r.get_json()["error"]


# ===========================================================================
# Excel import (audit C-05, L-03, L-06)
# ===========================================================================

class TestImportRoute:
    def _rows(self, *plates):
        return [
            {"vehicle": p, "sequence": i + 1, "station_code": f"S{i+1}",
             "station_name": f"Stop {i+1}", "lat": 10.8, "lng": 106.6}
            for i, p in enumerate(plates)
        ]

    def _vehicle_count(self, db):
        conn = sqlite3.connect(db)
        try:
            return conn.execute("SELECT COUNT(*) FROM vehicles").fetchone()[0]
        finally:
            conn.close()

    def test_plate_variants_collapse_to_one_assignment(self, client, db):
        """C-05 + L-03."""
        before = self._vehicle_count(db)
        plan_id = plan_service.create_plan(db, "Import", "2026-07-31")

        r = client.post("/api/plans/import/save", json={
            "plan_id": plan_id,
            "rows": self._rows("50E-18463", "50E18463", "50E 18463", "18463"),
        })

        assert r.status_code == 201
        assert r.get_json()["assignments_created"] == 1
        assert self._vehicle_count(db) == before, "import created duplicate vehicles"

    def test_unknown_vehicle_is_rejected_and_writes_nothing(self, client, db):
        before = self._vehicle_count(db)
        plan_id = plan_service.create_plan(db, "Import", "2026-07-31")

        r = client.post("/api/plans/import/save",
                             json={"plan_id": plan_id, "rows": self._rows("99Z-00000")})

        assert r.status_code == 409
        assert r.get_json()["unknown_vehicles"] == ["99Z-00000"]
        assert self._vehicle_count(db) == before
        assert plan_service.get_plan(db, plan_id)["status"] == "draft"

    def test_no_flag_can_make_the_import_create_a_vehicle(self, client, db):
        before = self._vehicle_count(db)
        plan_id = plan_service.create_plan(db, "Import", "2026-07-31")
        r = client.post("/api/plans/import/save", json={
            "plan_id": plan_id,
            "rows": self._rows("99Z-00000"),
            "create_missing_vehicles": True,   # inert
        })
        assert r.status_code == 409
        assert self._vehicle_count(db) == before

    def test_plan_is_confirmed_once_on_success(self, client, db):
        """L-06: the status UPDATE sat inside the per-vehicle loop."""
        plan_id = plan_service.create_plan(db, "Import", "2026-07-31")
        r = client.post("/api/plans/import/save",
                             json={"plan_id": plan_id, "rows": self._rows("50E-18463")})
        assert r.get_json()["plan_confirmed"] is True
        assert plan_service.get_plan(db, plan_id)["status"] == "confirmed"

    def test_import_requires_plan_id_and_rows(self, client):
        assert client.post("/api/plans/import/save", json={}).status_code == 400


# ===========================================================================
# Image upload (audit S-04, S-05, C-08)
# ===========================================================================

def _upload(client, stop_id, filename, content=b"binary-image-data", category="delivery"):
    return client.post(
        f"/api/stops/{stop_id}/images",
        data={"file": (io.BytesIO(content), filename), "category": category},
        content_type="multipart/form-data",
    )


class TestImageUpload:
    @pytest.mark.parametrize("filename", ["photo.jpg", "photo.JPG", "photo.png", "photo.webp"])
    def test_image_types_are_accepted(self, client, plan, filename):
        assert _upload(client, plan["stop_ids"][0], filename).status_code == 201

    @pytest.mark.parametrize("filename", ["clip.mp4", "clip.MP4", "clip.mov", "clip.webm"])
    def test_video_types_are_accepted(self, client, plan, filename):
        """Video evidence, 2026-08-15. Drivers shoot proof as video and the
        allow-list previously had nowhere to put it."""
        assert _upload(client, plan["stop_ids"][0], filename).status_code == 201

    @pytest.mark.parametrize("filename", [
        "payload.html", "payload.svg", "shell.php", "notes.txt", "noext",
        # Widening for video must not have widened past the three formats
        # chosen — these are video too, and still have to be refused.
        "clip.avi", "clip.mkv", "clip.3gp",
    ])
    def test_dangerous_types_are_rejected(self, client, plan, filename):
        """S-05: send_file infers Content-Type from the extension, so an
        uploaded .html was served as text/html from the app's own origin."""
        resp = _upload(client, plan["stop_ids"][0], filename)
        assert resp.status_code == 400
        assert "Unsupported file type" in resp.get_json()["error"]

    def test_oversized_upload_is_rejected(self, client, plan):
        big = b"x" * (11 * 1024 * 1024)
        resp = _upload(client, plan["stop_ids"][0], "huge.jpg", content=big)
        assert resp.status_code == 400
        assert "limit" in resp.get_json()["error"]

    def test_the_same_payload_passes_as_video(self, client, plan):
        """Per-kind caps, through a real request rather than the service alone.

        11 MB is over the image limit and under the video one, so this is the
        one assertion that fails if the caps ever collapse back into one.
        """
        from services.delivery import image_service

        big = b"x" * (11 * 1024 * 1024)
        resp = _upload(client, plan["stop_ids"][0], "long.mp4", content=big)
        assert resp.status_code == 201

        img = image_service.get_image(client.application.config["DB_PATH"],
                                      resp.get_json()["id"])
        assert img["media_kind"] == "video"
        (image_service.DATA_ROOT / img["relative_path"]).unlink(missing_ok=True)

    def test_listing_reports_media_kind(self, client, plan):
        """The dashboard picks <img> vs <video> off this field, so an assembled
        response missing it renders every video as a broken image."""
        from services.delivery import image_service

        stop_id = plan["stop_ids"][0]
        _upload(client, stop_id, "a.jpg")
        _upload(client, stop_id, "b.mp4")

        listed = client.get(f"/api/stops/{stop_id}/images")
        assert listed.status_code == 200
        kinds = {i["original_filename"]: i["media_kind"] for i in listed.get_json()}
        assert kinds == {"a.jpg": "image", "b.mp4": "video"}

        db_path = client.application.config["DB_PATH"]
        for i in image_service.list_images(db_path, stop_id):
            (image_service.DATA_ROOT / i["relative_path"]).unlink(missing_ok=True)

    def test_empty_upload_is_rejected(self, client, plan):
        assert _upload(client, plan["stop_ids"][0], "empty.jpg", content=b"").status_code == 400

    def test_traversal_in_category_cannot_escape(self, client, plan):
        """S-04: category and station_code were interpolated into the upload
        path, so `../../../static/js` wrote into served static files."""
        from services.delivery import image_service

        resp = _upload(client, plan["stop_ids"][0], "ok.jpg",
                       category="../../../static/js")
        assert resp.status_code == 201

        img = image_service.get_image(client.application.config["DB_PATH"],
                                      resp.get_json()["id"])
        full = (image_service.DATA_ROOT / img["relative_path"]).resolve()
        assert full.is_relative_to(image_service.UPLOAD_ROOT.resolve())
        full.unlink(missing_ok=True)

    def test_two_uploads_in_the_same_second_both_survive(self, client, plan):
        """C-08: filenames were `{unix_seconds}{ext}`, so the second photo
        silently overwrote the first and both rows pointed at one file."""
        from services.delivery import image_service

        stop_id = plan["stop_ids"][0]
        first = _upload(client, stop_id, "a.jpg", content=b"first")
        second = _upload(client, stop_id, "b.jpg", content=b"second")
        assert first.status_code == second.status_code == 201

        db_path = client.application.config["DB_PATH"]
        images = image_service.list_images(db_path, stop_id)
        paths = {i["relative_path"] for i in images}
        assert len(paths) == 2, "one upload overwrote the other"

        for i in images:
            (image_service.DATA_ROOT / i["relative_path"]).unlink(missing_ok=True)

    def test_upload_to_a_missing_stop_404s(self, client, db):
        assert _upload(client, 999999, "x.jpg").status_code == 404

    def test_upload_requires_a_file(self, client, plan):
        resp = client.post(f"/api/stops/{plan['stop_ids'][0]}/images",
                                data={}, content_type="multipart/form-data")
        assert resp.status_code == 400


class TestImageServing:
    def test_uploaded_image_can_be_fetched_back(self, client, plan):
        from services.delivery import image_service

        resp = _upload(client, plan["stop_ids"][0], "p.jpg", content=b"the-bytes")
        image_id = resp.get_json()["id"]

        served = client.get(f"/api/images/{image_id}/file")
        assert served.status_code == 200
        assert served.data == b"the-bytes"

        img = image_service.get_image(client.application.config["DB_PATH"], image_id)
        (image_service.DATA_ROOT / img["relative_path"]).unlink(missing_ok=True)

    def test_missing_image_404s(self, client):
        assert client.get("/api/images/999999/file").status_code == 404

    def test_video_is_served_with_a_video_content_type(self, client, plan):
        """send_file keys Content-Type off the extension. If it ever resolved
        to octet-stream the dashboard's <video> would refuse to play it."""
        from services.delivery import image_service

        resp = _upload(client, plan["stop_ids"][0], "clip.mp4", content=b"video-bytes")
        image_id = resp.get_json()["id"]

        served = client.get(f"/api/images/{image_id}/file")
        assert served.status_code == 200
        assert served.data == b"video-bytes"
        assert served.mimetype == "video/mp4"

        img = image_service.get_image(client.application.config["DB_PATH"], image_id)
        (image_service.DATA_ROOT / img["relative_path"]).unlink(missing_ok=True)

    def test_video_supports_range_requests(self, client, plan):
        """A browser seeking in a <video> issues a Range request and treats a
        200-with-whole-body as non-seekable. send_file's conditional=True
        default is what makes this work — see the comment in serve_image."""
        from services.delivery import image_service

        resp = _upload(client, plan["stop_ids"][0], "clip.mp4", content=b"0123456789")
        image_id = resp.get_json()["id"]

        served = client.get(f"/api/images/{image_id}/file",
                            headers={"Range": "bytes=2-5"})
        assert served.status_code == 206
        assert served.data == b"2345"

        img = image_service.get_image(client.application.config["DB_PATH"], image_id)
        (image_service.DATA_ROOT / img["relative_path"]).unlink(missing_ok=True)


class TestEvidenceRemoval:
    """Deleting mis-uploaded evidence, 2026-08-15.

    The endpoint predates this; what was missing was any caller. Dispatch runs
    many vehicles at once, evidence lands on the wrong stop, and the dashboard
    gallery was read-only, so there was no way to correct it.
    """

    def test_delete_removes_row_and_file(self, client, plan):
        from services.delivery import image_service

        resp = _upload(client, plan["stop_ids"][0], "wrong-stop.jpg")
        image_id = resp.get_json()["id"]

        db_path = client.application.config["DB_PATH"]
        path = image_service.DATA_ROOT / image_service.get_image(db_path, image_id)["relative_path"]
        assert path.exists()

        assert client.delete(f"/api/images/{image_id}").status_code == 200
        assert not path.exists()
        assert image_service.get_image(db_path, image_id) is None

    def test_delete_is_idempotent_enough_to_404(self, client, plan):
        resp = _upload(client, plan["stop_ids"][0], "gone.jpg")
        image_id = resp.get_json()["id"]

        assert client.delete(f"/api/images/{image_id}").status_code == 200
        # A dispatcher double-tapping the remove button must get a clean 404,
        # not a 500 from unlinking a file that is already gone.
        assert client.delete(f"/api/images/{image_id}").status_code == 404

    def test_reupload_after_delete_lands_cleanly(self, client, plan):
        """The correction workflow end to end: wrong stop, delete, re-upload
        to the right one."""
        from services.delivery import image_service

        wrong, right = plan["stop_ids"][0], plan["stop_ids"][1]
        image_id = _upload(client, wrong, "evidence.jpg").get_json()["id"]
        assert client.delete(f"/api/images/{image_id}").status_code == 200

        assert _upload(client, right, "evidence.jpg").status_code == 201

        db_path = client.application.config["DB_PATH"]
        assert image_service.list_images(db_path, wrong) == []
        moved = image_service.list_images(db_path, right)
        assert len(moved) == 1
        (image_service.DATA_ROOT / moved[0]["relative_path"]).unlink(missing_ok=True)


# ===========================================================================
# CRUD + validation surface
# ===========================================================================

class TestPlanCrud:
    def test_create_requires_name_and_date(self, client):
        assert client.post("/api/plans", json={"plan_date": "2026-07-31"}).status_code == 400
        assert client.post("/api/plans", json={"plan_name": "x"}).status_code == 400

    def test_create_then_fetch(self, client, db):
        plan_id = client.post(
            "/api/plans", json={"plan_name": "P", "plan_date": "2026-07-31"}
        ).get_json()["id"]
        assert client.get(f"/api/plans/{plan_id}").get_json()["plan_name"] == "P"

    def test_missing_plan_404s(self, client):
        assert client.get("/api/plans/999999").status_code == 404

    def test_delete_cascades_to_stops(self, client, db, plan):
        assert client.delete(f"/api/plans/{plan['plan_id']}").status_code == 200
        assert plan_service.list_stops(db, plan["assignment_id"]) == []

    def test_batch_delete_requires_ids(self, client):
        assert client.post("/api/plans/batch-delete", json={"plan_ids": []}).status_code == 400


class TestStopCrud:
    def test_list_requires_assignment_id(self, client):
        assert client.get("/api/stops").status_code == 400

    def test_get_stop_includes_images(self, client, plan):
        body = client.get(f"/api/stops/{plan['stop_ids'][0]}").get_json()
        assert body["images"] == []

    def test_missing_stop_404s(self, client):
        assert client.get("/api/stops/999999").status_code == 404

    def test_create_requires_assignment_id(self, client):
        assert client.post("/api/stops", json={"station_name": "x"}).status_code == 400

    def test_insert_temp_stop_reopens_a_completed_plan(self, client, db, plan):
        for stop_id in plan["stop_ids"]:
            client.post(f"/api/stops/{stop_id}/skip", json={"reason": "x"})
        assert plan_service.get_plan(db, plan["plan_id"])["status"] == "completed"

        r = client.post("/api/stops/insert", json={
            "assignment_id": plan["assignment_id"], "after_sequence": 1,
            "station_name": "Urgent", "lat": 10.9, "lng": 106.7,
        })
        assert r.status_code == 201
        assert plan_service.get_plan(db, plan["plan_id"])["status"] == "executing"
