"""End-of-day collection of proof photos into the operator's folder shape.

The photos are already on disk, but organised the way they were *written* —
``DeliveryPlans/YYYY/MM/DD/plate/station/category/`` — which is convenient
for uploading and useless for handing over. What the operator has always
built by hand looks like this::

    2_8_BacLieuGiaRai_CanThoOMon/
      HuynhQuocTrong_79791/            driver, then the plate's 5-digit serial
        HinhNhanHang_01_08/            loading photos, taken the day before
        HinhGiaoHang_02_08/
          CTOM19/                      station code
            ...unload and door photos
      HinhThungTrong/                  empty containers, one per driver
      manifest.xlsx

So this module rebuilds that shape into a ZIP rather than changing where
uploads land — moving the write path would strand every row already stored.

Every folder in that tree is created whether or not anything goes in it: the
stop folders are a checklist, and an empty CTOM19/ is how the operator sees
that nobody photographed CTOM19.

Two of those folders hold photos that never pass through a stop: the loading
shots from the evening before, and the empty-container shot at the end. They
are handed over during the export itself and live in ``delivery_day_images``.
"""
import io
import logging
import unicodedata
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Optional

from app.db import DatabaseManager
from services.plate_utils import normalize_plate

from . import image_service

logger = logging.getLogger(__name__)

#: Categories handed over at export time rather than at a stop.
DAY_CATEGORIES = ("loading", "empty_container")

#: Folder names, as the operator writes them.
FOLDER_LOADING = "HinhNhanHang"
FOLDER_DELIVERY = "HinhGiaoHang"
FOLDER_EMPTY_CONTAINER = "HinhThungTrong"

#: Where a day photo goes when its ``label`` is blank. Every ``loading`` row
#: written before the driver picker existed is in that state, so this is the
#: normal case for historical dates, not an error path.
FOLDER_UNKNOWN_DRIVER = "KhongRoTaiXe"

#: Which stop-photo categories belong in the delivery folder. Anything else
#: a stop happens to carry (an "extra", a mistyped category) is left out of
#: the handover rather than silently filed as proof of something.
STOP_EXPORT_CATEGORIES = ("unload", "door")


def strip_accents(value: str) -> str:
    """``Huỳnh Quốc Trọng`` → ``HuynhQuocTrong``.

    Vietnamese diacritics decompose under NFD and can then be dropped, with
    one exception: ``đ``/``Đ`` is a distinct letter rather than d-plus-mark,
    so NFD leaves it alone and it has to be mapped by hand. Missing that is
    the classic way this kind of function produces ``NguyʼnVn`` instead of
    ``NguyenVan``.

    Everything outside A-Z0-9 is dropped rather than replaced, because the
    result is a single folder name and the operator's existing folders run
    the words together.
    """
    text = (value or "").replace("đ", "d").replace("Đ", "D")
    decomposed = unicodedata.normalize("NFD", text)
    without_marks = "".join(ch for ch in decomposed if not unicodedata.combining(ch))
    return "".join(ch for ch in without_marks if ch.isalnum())


def driver_folder_name(driver_name: str, plate: str) -> str:
    """``HuynhQuocTrong_79791`` — driver, then the plate's 5-digit serial.

    ``normalize_plate`` is the same reduction the GPS matching uses, so the
    number here is the one the rest of the system already agrees on.
    """
    name = strip_accents(driver_name) or "KhongRoTaiXe"
    serial = normalize_plate(plate) or "00000"
    return f"{name}_{serial}"


def _ddmm(value) -> str:
    """``2026-08-02`` → ``02_08``. Zero-padded, matching the operator's
    subfolder names (the top-level folder is typed by hand and is not)."""
    day = _parse_day(value)
    return f"{day.day:02d}_{day.month:02d}" if day else "00_00"


def _parse_day(value) -> Optional[datetime]:
    if not value:
        return None
    try:
        return datetime.strptime(str(value)[:10], "%Y-%m-%d")
    except (TypeError, ValueError):
        return None


def _safe(value: str, fallback: str) -> str:
    return image_service._safe_path_segment(value, fallback)


# ── Day-scoped photos ────────────────────────────────────────────────

def day_image_folder(day_date: str, category: str) -> Path:
    day = _parse_day(day_date) or datetime.now()
    folder = (image_service.UPLOAD_ROOT / "day" / f"{day.year}"
              / f"{day.month:02d}" / f"{day.day:02d}"
              / _safe(category, "extra"))
    resolved = folder.resolve()
    if not resolved.is_relative_to(image_service.UPLOAD_ROOT.resolve()):
        raise image_service.UploadRejected("Invalid upload destination.")
    resolved.mkdir(parents=True, exist_ok=True)
    return resolved


def add_day_image(db_path: str, day_date: str, category: str, file_storage,
                  label: str = "") -> int:
    """Store one loading / empty-container photo for a day.

    One file per request on purpose. The whole-request cap is 25 MB
    (``config.MAX_UPLOAD_MB``), so a day's loading photos could never have
    been handed over in a single multipart POST — and uploading them
    individually also means a failed ZIP download does not throw away
    everything just uploaded.
    """
    import uuid

    if category not in DAY_CATEGORIES:
        raise image_service.UploadRejected(
            f"Unknown category '{category}'. Expected one of: {', '.join(DAY_CATEGORIES)}."
        )
    if not _parse_day(day_date):
        raise image_service.UploadRejected("A valid date (YYYY-MM-DD) is required.")

    ext = image_service._validate_upload(file_storage)
    folder = day_image_folder(day_date, category)
    original_name = file_storage.filename or "photo"
    filename = f"{int(datetime.now().timestamp())}-{uuid.uuid4().hex[:8]}{ext}"
    file_path = folder / filename
    file_storage.save(str(file_path))

    relative = str(file_path.relative_to(image_service.DATA_ROOT))

    with DatabaseManager(db_path).connect() as conn:
        c = conn.cursor()
        try:
            c.execute("""
                INSERT INTO delivery_day_images
                    (day_date, category, label, filename, relative_path, original_filename)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (day_date[:10], category, (label or "").strip(), filename,
                  relative, original_name))
        except Exception:
            # Same ordering as image_service.upload_image: the file only
            # survives if its row does, so a failed insert cannot leave an
            # orphan on a disk nobody is watching.
            if file_path.exists():
                file_path.unlink()
            raise
        return c.lastrowid


def list_day_images(db_path: str, day_date: str, category: Optional[str] = None) -> list:
    with DatabaseManager(db_path).connect() as conn:
        c = conn.cursor()
        if category:
            c.execute("""
                SELECT * FROM delivery_day_images
                WHERE day_date = ? AND category = ? ORDER BY id
            """, (day_date[:10], category))
        else:
            c.execute("""
                SELECT * FROM delivery_day_images WHERE day_date = ? ORDER BY id
            """, (day_date[:10],))
        return [image_service._with_kind(dict(r)) for r in c.fetchall()]


def delete_day_image(db_path: str, image_id: int) -> bool:
    with DatabaseManager(db_path).connect() as conn:
        c = conn.cursor()
        c.execute("SELECT relative_path FROM delivery_day_images WHERE id = ?", (image_id,))
        row = c.fetchone()
        if row:
            full_path = image_service.DATA_ROOT / row["relative_path"]
            try:
                if full_path.exists():
                    full_path.unlink()
            except Exception as e:
                logger.warning("Failed to delete file %s: %s", full_path, e)
        c.execute("DELETE FROM delivery_day_images WHERE id = ?", (image_id,))
        return c.rowcount > 0


# ── What the day contains ────────────────────────────────────────────

def day_summary(db_path: str, day_date: str) -> dict:
    """Everything the export screen and the manifest need, in one read.

    Includes stops whose proof is incomplete: the point of showing this
    before the download is that a missing photo can still be chased while
    the driver is reachable.
    """
    day = (day_date or "")[:10]
    with DatabaseManager(db_path).connect() as conn:
        c = conn.cursor()
        c.execute("""
            SELECT
                s.id AS stop_id,
                s.station_code,
                s.station_name,
                e.status AS execution_status,
                va.id AS assignment_id,
                v.plate_number,
                COALESCE(NULLIF(va.driver_name_override, ''), NULLIF(d.name, ''), v.current_driver, '') AS driver_name,
                dp.plan_name
            FROM delivery_plan_stops s
            JOIN vehicle_assignments va ON va.id = s.vehicle_assignment_id
            JOIN delivery_plans dp ON dp.id = va.plan_id
            LEFT JOIN stop_executions e ON e.stop_id = s.id
            LEFT JOIN vehicles v ON v.id = va.vehicle_id
            LEFT JOIN drivers d ON d.id = va.driver_id
            WHERE dp.plan_date = ?
            ORDER BY va.id, COALESCE(e.execution_sequence, s.planned_sequence)
        """, (day,))
        stops = [dict(r) for r in c.fetchall()]

        if stops:
            placeholders = ",".join("?" for _ in stops)
            c.execute(f"""
                SELECT stop_id, category, COUNT(*) AS n
                FROM delivery_stop_images
                WHERE stop_id IN ({placeholders})
                GROUP BY stop_id, category
            """, [s["stop_id"] for s in stops])
            by_stop = {}
            for row in c.fetchall():
                by_stop.setdefault(row["stop_id"], {})[row["category"]] = row["n"]
        else:
            by_stop = {}

        # The override reason, where one was used. This is the only record
        # that a completion was waived, so it belongs in the handover.
        c.execute("""
            SELECT stop_id, reason FROM stop_status_events
            WHERE to_status = 'completed' AND action = 'advance' AND reason != ''
            ORDER BY id
        """)
        overrides = {r["stop_id"]: r["reason"] for r in c.fetchall()}

    drivers = {}
    for stop in stops:
        counts = by_stop.get(stop["stop_id"], {})
        stop["photo_counts"] = counts
        stop["missing"] = [cat for cat in STOP_EXPORT_CATEGORIES if not counts.get(cat)]
        stop["override_reason"] = overrides.get(stop["stop_id"], "")
        stop["folder"] = driver_folder_name(stop["driver_name"], stop["plate_number"])
        drivers.setdefault(stop["folder"], {
            "folder": stop["folder"],
            "driver_name": stop["driver_name"],
            "plate_number": stop["plate_number"],
            "stops": [],
        })["stops"].append(stop)

    return {
        "date": day,
        "drivers": list(drivers.values()),
        "stop_count": len(stops),
        "incomplete_count": sum(1 for s in stops if s["missing"]),
        "day_images": {
            cat: list_day_images(db_path, day, cat) for cat in DAY_CATEGORIES
        },
    }


# ── The ZIP ──────────────────────────────────────────────────────────

MANIFEST_HEADERS = [
    "driver", "plate", "station_code", "station_name", "status",
    "unload_photos", "door_photos", "missing", "override_reason",
]


def _manifest_rows(summary: dict):
    for driver in summary["drivers"]:
        for stop in driver["stops"]:
            counts = stop["photo_counts"]
            yield [
                driver["driver_name"], driver["plate_number"],
                stop["station_code"], stop["station_name"],
                stop["execution_status"] or "",
                counts.get("unload", 0), counts.get("door", 0),
                " ".join(stop["missing"]),
                stop["override_reason"],
            ]


def _manifest_xlsx(summary: dict) -> bytes:
    """The manifest as a real spreadsheet.

    It was a CSV until 2026-08-10 and the driver and station names came out of
    Excel as mojibake — ``Huỳnh Quốc Trọng`` as ``Huá»³nh``. Nothing was wrong
    with the file: ZIP entries were written UTF-8, but Excel opens a .csv in
    the machine's ANSI codepage unless the bytes start with a BOM, and on a
    Vietnamese Windows that is CP1258. A BOM would have fixed it; .xlsx is
    proof against the whole class of question, since the encoding is declared
    inside the format rather than guessed at by the reader.

    openpyxl is already a dependency (the plan importer reads .xlsx).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "manifest"

    ws.append(MANIFEST_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    rows = list(_manifest_rows(summary))
    for row in rows:
        ws.append(row)

    # Width from the longest value actually present. Vietnamese names are long
    # and a column of ##### is its own kind of unreadable.
    for i, header in enumerate(MANIFEST_HEADERS, start=1):
        longest = max([len(str(header))] + [len(str(r[i - 1])) for r in rows])
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = \
            min(longest + 2, 40)

    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_day_zip(db_path: str, day_date: str, folder_name: str,
                  loading_date: Optional[str] = None) -> io.BytesIO:
    """Assemble the handover ZIP for one delivery day.

    ``folder_name`` is typed by the operator — it carries the route names,
    which exist nowhere in the data. It is sanitized to a single safe path
    component before use.

    ``loading_date`` stamps the HinhNhanHang folder and defaults to the day
    before the delivery, which is when loading happens.

    Missing files are skipped with a warning rather than aborting: a ZIP of
    everything that survived is more use at 6pm than an error, and the
    manifest still records what was expected.
    """
    summary = day_summary(db_path, day_date)
    root = _safe(folder_name, f"export_{_ddmm(day_date)}")

    day = _parse_day(day_date)
    if loading_date:
        loading_ddmm = _ddmm(loading_date)
    elif day:
        from datetime import timedelta
        loading_ddmm = _ddmm((day - timedelta(days=1)).strftime("%Y-%m-%d"))
    else:
        loading_ddmm = "00_00"

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        written = set()
        dirs_written = set()

        def add_dir(arcname: str):
            """Create a folder in the ZIP whether or not anything goes in it.

            The operator's tree has a folder per stop and it is a checklist:
            an empty CTBTX0/ is how they see that nobody photographed CTBTX0.
            Before 2026-08-10 a stop with no photos simply had no folder, so a
            missed stop and a stop that does not exist looked identical.

            A ZIP directory is a zero-length entry whose name ends in "/" and
            whose external_attr carries the directory bit — both the Unix mode
            and the MS-DOS flag, because Windows Explorer reads the latter and
            would otherwise unpack these as empty files.
            """
            if not arcname.endswith("/"):
                arcname += "/"
            if arcname in dirs_written:
                return
            dirs_written.add(arcname)
            info = zipfile.ZipInfo(arcname)
            info.external_attr = (0o40755 << 16) | 0x10
            zf.writestr(info, b"")

        def add(arcname: str, source: Path):
            if not source.exists():
                logger.warning("Export: missing file on disk, skipped: %s", source)
                return
            # Two photos can share an original filename; ZIP happily stores
            # duplicates and the operator's file manager silently keeps one.
            name = arcname
            n = 2
            while name in written:
                stem, dot, ext = arcname.rpartition(".")
                name = f"{stem}_{n}{dot}{ext}" if dot else f"{arcname}_{n}"
                n += 1
            written.add(name)
            zf.write(source, name)

        # Both photo folders live under the driver, matching the tree the
        # operator handed over on 2026-08-10:
        #
        #   8_8/
        #     TranHoangQuan_79107/
        #       HinhNhanHang_07_08/      loading, loose — see below
        #       HinhGiaoHang_08_08/
        #         CTVT88/                one per stop, created empty if unshot
        #     HinhThungTrong/
        #     manifest.xlsx
        #
        # HinhNhanHang sat at the top level for one day between the per-driver
        # split and this; it was never handed over in that shape.
        delivery_folder = f"{FOLDER_DELIVERY}_{_ddmm(day_date)}"
        loading_folder = f"{FOLDER_LOADING}_{loading_ddmm}"

        for driver in summary["drivers"]:
            # Both folders exist for every driver on the plan, photos or not.
            add_dir(f"{root}/{driver['folder']}/{loading_folder}")
            add_dir(f"{root}/{driver['folder']}/{delivery_folder}")

            for stop in driver["stops"]:
                station = _safe(stop["station_code"], "KhongRoTram")
                stop_dir = f"{root}/{driver['folder']}/{delivery_folder}/{station}"
                add_dir(stop_dir)
                for image in image_service.list_images(db_path, stop["stop_id"]):
                    if image["category"] not in STOP_EXPORT_CATEGORIES:
                        continue
                    source = image_service.DATA_ROOT / image["relative_path"]
                    add(f"{stop_dir}/{image['category']}_{image['filename']}", source)

        # Loading photos: inside the driver's folder, but *not* split by stop —
        # the operator's call. Loading happens as a truck is filled, one pass,
        # and asking for a station on every shot would be a tap per photo for a
        # distinction nobody reads afterwards.
        #
        # The upload picker stores the *folder* name in `label` (see
        # renderDrivers in delivery-export.js), not the driver's name, so a
        # photo lands beside that driver's HinhGiaoHang without a lookup.
        # Deriving it from a driver_name here instead would be ambiguous the
        # moment one driver runs two trucks: two folders, one name.
        for image in summary["day_images"]["loading"]:
            driver_dir = _safe(image["label"], FOLDER_UNKNOWN_DRIVER)
            source = image_service.DATA_ROOT / image["relative_path"]
            add(f"{root}/{driver_dir}/{loading_folder}/"
                f"{image['original_filename'] or image['filename']}",
                source)

        # Empty containers: still one flat folder, but the driver's name goes
        # into the filename — one photo per driver, and a folder of anonymous
        # truck interiors tells you nothing.
        add_dir(f"{root}/{FOLDER_EMPTY_CONTAINER}")
        for image in summary["day_images"]["empty_container"]:
            label = strip_accents(image["label"]) or FOLDER_UNKNOWN_DRIVER
            source = image_service.DATA_ROOT / image["relative_path"]
            add(f"{root}/{FOLDER_EMPTY_CONTAINER}/{label}_{image['filename']}", source)

        zf.writestr(f"{root}/manifest.xlsx", _manifest_xlsx(summary))

    buffer.seek(0)
    return buffer
